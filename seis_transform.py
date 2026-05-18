"""Transform SEIS wide-format report into Aeries narrow-format import file.

SEIS exports one row per student with many setting columns.
Aeries STS table expects one row per student per setting.
This unpivots the data, filters out settings that already exist
in Aeries, and outputs only new settings for import.
"""

import csv
import openpyxl
import os


def read_seis_settings(seis_report_path):
    """Read SEIS report into a dict of SSID -> set of setting codes.

    Args:
        seis_report_path: Path to the SEIS TOMS report xlsx

    Returns:
        Dict mapping SSID strings to sets of setting code strings
    """
    wb = openpyxl.load_workbook(seis_report_path, read_only=True)
    ws = wb.active

    students = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # Skip header
        ssid = str(row[0]).strip() if row[0] else ''
        if not ssid:
            continue
        settings = set()
        for j in range(1, len(row)):
            val = str(row[j]).strip() if row[j] else ''
            if val:
                settings.add(val)
        students[ssid] = settings

    wb.close()
    return students


def read_aeries_settings(aeries_csv_path):
    """Read Aeries export CSV into a dict of SSID -> set of setting codes.

    Args:
        aeries_csv_path: Path to the Aeries test settings CSV

    Returns:
        Dict mapping SSID strings to sets of setting code strings
    """
    students = {}
    with open(aeries_csv_path, newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row in reader:
            ssid = row[0].strip()
            if not ssid:
                continue
            settings = set()
            for val in row[1:]:
                val = val.strip()
                if val:
                    settings.add(val)
            students[ssid] = settings
    return students


def transform_seis_to_aeries(seis_report_path, aeries_csv_path=None, output_path=None):
    """Transform SEIS TOMS report into Aeries STS import format.

    Reads the wide-format SEIS xlsx, compares against current Aeries
    settings if provided, unpivots only NEW settings into rows, and
    outputs a CSV for Aeries import.

    Args:
        seis_report_path: Path to the SEIS TOMS report xlsx
        aeries_csv_path: Path to current Aeries export CSV (for filtering)
        output_path: Path for the output CSV (defaults to same folder)

    Returns:
        Path to the output CSV, or None if no new settings found
    """
    if output_path is None:
        output_dir = os.path.dirname(seis_report_path)
        output_path = os.path.join(output_dir, "seis_aeries_import.csv")

    # Read SEIS settings
    print(f"Reading SEIS report: {seis_report_path}")
    seis_students = read_seis_settings(seis_report_path)
    total_seis_settings = sum(len(s) for s in seis_students.values())
    print(f"  {len(seis_students)} students, {total_seis_settings} total settings in SEIS.")

    # Read current Aeries settings if provided
    if aeries_csv_path:
        print(f"Reading current Aeries settings: {aeries_csv_path}")
        aeries_students = read_aeries_settings(aeries_csv_path)
        total_aeries_settings = sum(len(s) for s in aeries_students.values())
        print(f"  {len(aeries_students)} students, {total_aeries_settings} total settings in Aeries.")
    else:
        aeries_students = {}
        print("  No Aeries export provided — importing all SEIS settings.")

    # Find new settings: in SEIS but not in Aeries
    output_rows = []
    new_students = 0
    for ssid, seis_settings in seis_students.items():
        aeries_settings = aeries_students.get(ssid, set())
        new_settings = seis_settings - aeries_settings

        if ssid not in aeries_students and seis_settings:
            new_students += 1

        for setting in new_settings:
            output_rows.append([ssid, setting])

    print(f"\n  Comparison results:")
    print(f"    New students (in SEIS, not in Aeries): {new_students}")
    print(f"    New settings to import: {len(output_rows)}")
    print(f"    Settings already in Aeries (skipped): {total_seis_settings - len(output_rows)}")

    if not output_rows:
        print("  No new settings to import.")
        return None

    # Write CSV
    with open(output_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['CID', 'CD'])
        for row in output_rows:
            writer.writerow(row)

    print(f"  Aeries import file saved: {output_path}")
    print(f"  {len(output_rows)} new settings ready to import.")
    return output_path

def generate_review_report(seis_report_path, aeries_csv_path, output_path=None):
    """Generate a human-readable report of new settings from SEIS not yet in Aeries.

    Includes student names from the SEIS report for easy review.

    Args:
        seis_report_path: Path to the SEIS TOMS report xlsx
        aeries_csv_path: Path to current Aeries export CSV
        output_path: Path for the output CSV

    Returns:
        Path to the review report
    """
    if output_path is None:
        output_dir = os.path.dirname(seis_report_path)
        output_path = os.path.join(output_dir, "seis_review_report.csv")

    # Read SEIS report with headers and names
    wb = openpyxl.load_workbook(seis_report_path, read_only=True)
    ws = wb.active

    headers = []
    seis_data = {}  # ssid -> {settings: set, name: str, row_data: dict}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h) if h else '' for h in row]
            continue
        ssid = str(row[0]).strip() if row[0] else ''
        if not ssid:
            continue

        settings = {}
        for j in range(1, len(row)):
            val = str(row[j]).strip() if row[j] else ''
            if val:
                col_name = headers[j] if j < len(headers) else f'Column {j}'
                settings[val] = col_name

        seis_data[ssid] = {'settings': settings}

    wb.close()

    # Read current Aeries settings
    aeries_students = read_aeries_settings(aeries_csv_path)

    # Build the review report
    report_rows = []
    for ssid, data in seis_data.items():
        aeries_settings = aeries_students.get(ssid, set())
        is_new_student = ssid not in aeries_students

        for code, column_name in data['settings'].items():
            if code not in aeries_settings:
                report_rows.append({
                    'SSID': ssid,
                    'Status': 'NEW STUDENT' if is_new_student else 'New Setting',
                    'Setting Code': code,
                    'Setting Description': column_name.split('\n')[0].strip(),
                })

    if not report_rows:
        print("  No new settings to report.")
        return None

    # Write CSV
    with open(output_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['SSID', 'Status', 'Setting Code', 'Setting Description'])
        writer.writeheader()
        writer.writerows(report_rows)

    print(f"  Review report saved: {output_path}")
    print(f"  {len(report_rows)} new settings across {len(set(r['SSID'] for r in report_rows))} students.")

    return output_path